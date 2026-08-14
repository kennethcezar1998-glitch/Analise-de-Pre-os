"""
Bot de analise de lucro liquido.

Le as planilhas:
  - "Analise de lucro liquido.xlsx"          (vendas, uma linha por pedido/codigo auxiliar)
  - "PRODUTOS BASE.xlsx"                     (cadastro de produtos, fonte principal do Produto Pai e do nome)
  - "BASE DE PRODUTOS SITE SPORTBAY.xlsx"    (cadastro do site, usado so como reserva se o codigo
                                               nao for encontrado em PRODUTOS BASE)

E gera "Analise de lucro liquido - Resultado.xlsx" com duas abas:

  Aba "Sheet1" (mesmo formato da aba "Sheet1" da planilha "Analise de lucro liquido - Modelo"):
    uma linha por venda, com o Produto Pai ja resolvido na coluna A e as formulas
    de apoio (Qtd., Media, %) recalculadas dentro da propria planilha:
      A) pai            -> Produto Pai da venda
      B) Qtd.           -> =CONT.SE(A:A;A2)              (quantas vezes o pai aparece)
      C) pedido         -> coluna A da planilha de vendas
      D) cod. Auxiliar  -> coluna B da planilha de vendas
      E..N              -> colunas financeiras da planilha de vendas (C..L)
      O) % Lucro Liquido-> coluna M da planilha de vendas
      P) Media          -> =SOMASE(A:A;A2;O:O)
      Q) %              -> =P2/B2   (media do % lucro liquido do Produto Pai)

  Aba "Planilha1" (resultado agrupado por Produto Pai, uma linha por produto):
      A) pai            -> Produto Pai
      B) produto        -> nome do produto (PRODUTOS BASE, coluna H)
      C) qtd            -> quantas vezes aquele Produto Pai apareceu na planilha de vendas
      D) media de lucro -> media do "% Lucro Liquido" das vendas daquele Produto Pai

Uso:
    python "analisar_lucro_liquido.py"

Os 4 arquivos devem estar na mesma pasta deste script (a planilha "- Modelo" e ignorada,
ela serve apenas de referencia do formato esperado).
"""

import os
from collections import defaultdict, Counter

import openpyxl

PASTA = os.path.dirname(os.path.abspath(__file__))

ARQ_ANALISE = os.path.join(PASTA, "Analise de lucro liquido.xlsx")
ARQ_PRODUTOS_BASE = os.path.join(PASTA, "PRODUTOS BASE.xlsx")
ARQ_SPORTBAY = os.path.join(PASTA, "BASE DE PRODUTOS SITE SPORTBAY.xlsx")
ARQ_SAIDA = os.path.join(PASTA, "Analise de lucro liquido - Resultado.xlsx")

# indices (base 0) das colunas na planilha "Analise de lucro liquido.xlsx"
COL_ANALISE_PEDIDO = 0
COL_ANALISE_COD_AUXILIAR = 1
COL_ANALISE_PCT_LUCRO = 12  # % Lucro Liquido (ultima coluna)

COL_BASE_COD_AUXILIAR = 1  # coluna B
COL_BASE_PRODUTO_PAI = 3  # coluna D
COL_BASE_PRODUTO_NOME = 7  # coluna H

COL_SPORTBAY_COD_AUXILIAR = 1  # coluna B
COL_SPORTBAY_DESCRICAO = 11  # coluna L (Descricao do Produto)


def normalizar(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto if texto else None


def carregar_produtos_base(caminho):
    """Mapa codigo auxiliar -> (produto pai, nome do produto)."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    mapa = {}
    for linha in ws.iter_rows(min_row=2, values_only=True):
        codigo = normalizar(linha[COL_BASE_COD_AUXILIAR])
        if codigo is None:
            continue
        pai = normalizar(linha[COL_BASE_PRODUTO_PAI])
        nome = linha[COL_BASE_PRODUTO_NOME]
        mapa[codigo] = (pai, nome)
    wb.close()
    return mapa


def carregar_sportbay_reserva(caminho):
    """Mapa reserva codigo auxiliar -> (produto pai, nome), usado quando o
    codigo nao existe em PRODUTOS BASE. O Produto Pai nao vem pronto nessa
    planilha, entao e derivado do prefixo antes do ponto no codigo auxiliar."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    mapa = {}
    for linha in ws.iter_rows(min_row=2, values_only=True):
        codigo = normalizar(linha[COL_SPORTBAY_COD_AUXILIAR])
        if codigo is None:
            continue
        pai = codigo.split(".")[0]
        nome = linha[COL_SPORTBAY_DESCRICAO]
        mapa[codigo] = (pai, nome)
    wb.close()
    return mapa


def carregar_vendas(caminho):
    """Lista de linhas de venda (pedido, cod. auxiliar, 9 colunas financeiras, % lucro)."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active
    vendas = []
    for linha in ws.iter_rows(min_row=2, values_only=True):
        codigo = normalizar(linha[COL_ANALISE_COD_AUXILIAR])
        if codigo is None:
            continue  # ignora linha de total ou linhas em branco
        vendas.append(linha[:COL_ANALISE_PCT_LUCRO + 1])
    wb.close()
    return vendas


def resolver_pai_nome(codigo, produtos_base, sportbay, nao_encontrados):
    info = produtos_base.get(codigo) or sportbay.get(codigo)
    if info is None:
        nao_encontrados.add(codigo)
        return codigo, f"(nao encontrado) {codigo}"
    pai, nome = info
    if pai is None:
        pai = codigo
    return pai, nome


def montar_linhas_completas(vendas, produtos_base, sportbay):
    """Resolve o Produto Pai/nome de cada venda e devolve uma linha por venda:
    (pai, nome, pedido, cod_auxiliar, *colunas financeiras, pct_lucro)."""
    nao_encontrados = set()
    linhas = []
    for venda in vendas:
        codigo = normalizar(venda[COL_ANALISE_COD_AUXILIAR])
        pai, nome = resolver_pai_nome(codigo, produtos_base, sportbay, nao_encontrados)
        linhas.append((pai, nome, *venda))
    return linhas, nao_encontrados


def montar_resultado_agrupado(linhas_completas):
    grupos = defaultdict(lambda: {"qtd": 0, "soma_pct": 0.0, "nomes": Counter()})

    for pai, nome, *_resto, pct_lucro in linhas_completas:
        pct_lucro = pct_lucro or 0
        grupo = grupos[pai]
        grupo["qtd"] += 1
        grupo["soma_pct"] += pct_lucro
        grupo["nomes"][nome] += 1

    resultado = []
    for pai, dados in grupos.items():
        nome_mais_frequente = dados["nomes"].most_common(1)[0][0]
        media_lucro = dados["soma_pct"] / dados["qtd"]
        resultado.append((pai, nome_mais_frequente, dados["qtd"], media_lucro))

    resultado.sort(key=lambda item: (item[1] or ""))
    return resultado


def escrever_sheet1(wb, linhas_completas):
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([
        "pai", "Qtd.", "pedido", "cod. Auxiliar",
        "R$ Venda Total", "R$ Custo Formação", "R$ Impostos", "R$ Custo do Frete",
        "R$ Frete Cobrado", "R$ Comissão", "R$ Rebait", "R$ Despesas Fixas",
        "R$ Despesas Variáveis", "R$ Lucro Líquido", "% Lucro Líquido",
        "Media", "%",
    ])

    for i, (pai, _nome, pedido, cod_aux, *financeiro) in enumerate(linhas_completas, start=2):
        ws.append([
            pai,
            f"=COUNTIF(A:A,A{i})",
            pedido,
            cod_aux,
            *financeiro,
            f"=SUMIF(A:A,A{i},O:O)",
            f"=P{i}/B{i}",
        ])

    larguras = {
        "A": 14, "B": 8, "C": 14, "D": 16, "E": 14, "F": 14, "G": 12, "H": 14,
        "I": 14, "J": 12, "K": 10, "L": 14, "M": 16, "N": 14, "O": 14, "P": 12, "Q": 10,
    }
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura


def escrever_resultado_agrupado(wb, resultado):
    ws = wb.create_sheet("Planilha1")
    ws.append(["pai", "produto", "qtd", "media de lucro"])
    for pai, nome, qtd, media in resultado:
        ws.append([pai, nome, qtd, round(media, 4)])

    larguras = {"A": 16, "B": 60, "C": 8, "D": 16}
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura
    for celula in ws["D"][1:]:
        celula.number_format = "0.00"


def main():
    print("Lendo PRODUTOS BASE.xlsx...")
    produtos_base = carregar_produtos_base(ARQ_PRODUTOS_BASE)

    print("Lendo BASE DE PRODUTOS SITE SPORTBAY.xlsx (reserva)...")
    sportbay = carregar_sportbay_reserva(ARQ_SPORTBAY)

    print("Lendo Analise de lucro liquido.xlsx...")
    vendas = carregar_vendas(ARQ_ANALISE)

    linhas_completas, nao_encontrados = montar_linhas_completas(vendas, produtos_base, sportbay)
    resultado = montar_resultado_agrupado(linhas_completas)

    wb = openpyxl.Workbook()
    escrever_sheet1(wb, linhas_completas)
    escrever_resultado_agrupado(wb, resultado)
    wb.save(ARQ_SAIDA)

    print()
    print(f"Linhas de venda processadas: {len(vendas)}")
    print(f"Produtos Pai (linhas na aba Planilha1): {len(resultado)}")
    if nao_encontrados:
        print(f"ATENCAO: {len(nao_encontrados)} codigo(s) auxiliar(es) nao encontrados em nenhuma base:")
        for codigo in sorted(nao_encontrados):
            print(f"  - {codigo}")
    print(f"Arquivo gerado: {ARQ_SAIDA}")


if __name__ == "__main__":
    main()
