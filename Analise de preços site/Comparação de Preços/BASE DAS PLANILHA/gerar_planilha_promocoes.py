# -*- coding: utf-8 -*-
"""
Bot de geração da planilha de PROMOÇÕES.

Cruza os dados de 5 planilhas:
  - Base dos produtos para alteração.xlsx  -> lista de códigos PAI a processar
  - PRODUTOS BASE.xlsx                     -> descobre os produtos filhos de cada PAI
  - BASE DE PRODUTOS SITE SPORTBAY.xlsx    -> preço do site e data de fim da promoção
  - PROMOÇÕES - MODELO.xlsx                -> modelo de formatação da planilha de promoções
  - PREÇO - MODELO.xlsx                    -> modelo de formatação da planilha de divergências

Para cada código (pai ou filho), se o preço da coluna E de "Base dos produtos
para alteração" for MAIOR que o Preço Site (coluna Q) daquele código, há uma
divergência: o código não entra na planilha de promoções e vai, em vez disso,
para a planilha de divergências (com o preço "cheio" da coluna E).

Gera: PROMOÇÕES - GERADA.xlsx
Gera (somente se houver divergência): PREÇO - DIVERGÊNCIAS.xlsx
"""

import datetime as dt
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel

BASE_DIR = Path(__file__).resolve().parent

ARQ_ALTERACAO = BASE_DIR / "Base dos produtos para alteração.xlsx"
ARQ_PRODUTOS_BASE = BASE_DIR / "PRODUTOS BASE.xlsx"
ARQ_SITE_SPORTBAY = BASE_DIR / "BASE DE PRODUTOS SITE SPORTBAY.xlsx"
ARQ_MODELO = BASE_DIR / "PROMOÇÕES - MODELO.xlsx"
ARQ_SAIDA = BASE_DIR / "PROMOÇÕES - GERADA.xlsx"
ARQ_MODELO_DIVERGENCIA = BASE_DIR / "PREÇO - MODELO.xlsx"
ARQ_SAIDA_DIVERGENCIA = BASE_DIR / "PREÇO - DIVERGÊNCIAS.xlsx"

MINUTOS_ATRASO_HORA_INICIO = 10
HORA_FIM = dt.time(23, 59, 59)

# A coluna E de "Base dos produtos para alteração" traz o valor já com o
# desconto de 7% do Pix aplicado. Dividimos por 0,93 para obter o preço
# campanha "cheio" que, com os 7% de desconto no Pix, resulta nesse valor.
FATOR_DESCONTO_PIX = 0.93


def normalizar(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def ler_produtos_para_alterar():
    """Coluna A = código PAI, coluna E = preço (Sugestão de preço final, já com desconto Pix)."""
    wb = openpyxl.load_workbook(ARQ_ALTERACAO, read_only=True, data_only=True)
    ws = wb.active
    produtos = []
    for linha in ws.iter_rows(min_row=2, values_only=True):
        codigo_pai = normalizar(linha[0])
        if codigo_pai is None:
            continue
        preco_e = linha[4]
        produtos.append((codigo_pai, preco_e))
    wb.close()
    return produtos


def ler_produtos_base():
    """Coluna B = Código Auxiliar, coluna D = Produto PAI (só preenchida nos filhos)."""
    wb = openpyxl.load_workbook(ARQ_PRODUTOS_BASE, read_only=True, data_only=True)
    ws = wb.active
    codigos_existentes = set()
    filhos_por_pai = {}
    for linha in ws.iter_rows(min_row=2, values_only=True):
        codigo = normalizar(linha[1])
        pai = normalizar(linha[3])
        if codigo is None:
            continue
        codigos_existentes.add(codigo)
        if pai is not None:
            filhos_por_pai.setdefault(pai, []).append(codigo)
    wb.close()
    return codigos_existentes, filhos_por_pai


def ler_site_sportbay():
    """Coluna B = Código Auxiliar, coluna Q = Preço Site, coluna U = Fim da Promoção."""
    wb = openpyxl.load_workbook(ARQ_SITE_SPORTBAY, read_only=True, data_only=True)
    ws = wb.active
    preco_site_por_codigo = {}
    fim_promocao_por_codigo = {}
    for linha in ws.iter_rows(min_row=2, values_only=True):
        codigo = normalizar(linha[1])
        if codigo is None:
            continue
        preco_site_por_codigo[codigo] = linha[16]
        fim_promocao_por_codigo[codigo] = linha[20]
    wb.close()
    return preco_site_por_codigo, fim_promocao_por_codigo


def para_data_sem_hora(valor):
    # Algumas células de "Fim da Promoção" na planilha de origem não têm
    # formato de data aplicado e chegam como número serial do Excel.
    if isinstance(valor, (int, float)):
        try:
            valor = from_excel(valor)
        except (ValueError, OverflowError):
            return None
    if isinstance(valor, dt.datetime):
        return dt.datetime.combine(valor.date(), dt.time())
    if isinstance(valor, dt.date):
        return dt.datetime.combine(valor, dt.time())
    return None


def montar_linhas(produtos_alterar, codigos_existentes, filhos_por_pai,
                   preco_site_por_codigo, fim_promocao_por_codigo):
    agora = dt.datetime.now()
    data_inicio = dt.datetime.combine(agora.date(), dt.time())
    hora_inicio = (agora - dt.timedelta(minutes=MINUTOS_ATRASO_HORA_INICIO)).time().replace(microsecond=0)

    linhas = []
    linhas_divergencia = []
    avisos = []

    for codigo_pai, preco_e in produtos_alterar:
        if codigo_pai not in codigos_existentes:
            avisos.append(f"Código PAI '{codigo_pai}' não encontrado em PRODUTOS BASE.")

        grupo = [codigo_pai] + filhos_por_pai.get(codigo_pai, [])

        for codigo in grupo:
            if codigo not in preco_site_por_codigo:
                avisos.append(
                    f"Código '{codigo}' não encontrado em BASE DE PRODUTOS SITE SPORTBAY "
                    "- não incluído na planilha final."
                )
                continue

            preco_site = preco_site_por_codigo.get(codigo)

            if isinstance(preco_e, (int, float)) and isinstance(preco_site, (int, float)) \
                    and preco_e > preco_site:
                avisos.append(
                    f"Código '{codigo}': preço da coluna E ({preco_e}) é maior que o "
                    f"Preço Site ({preco_site}) - divergência, movido para "
                    f"'{ARQ_SAIDA_DIVERGENCIA.name}'."
                )
                linhas_divergencia.append({
                    "codigo": codigo,
                    "preco": preco_e,
                })
                continue

            fim_bruto = fim_promocao_por_codigo.get(codigo)
            preco_campanha = preco_e / FATOR_DESCONTO_PIX if isinstance(preco_e, (int, float)) else None

            linhas.append({
                "codigo": codigo,
                "preco_site": preco_site,
                "preco_campanha": preco_campanha,
                "data_inicio": data_inicio,
                "hora_inicio": hora_inicio,
                "data_fim": para_data_sem_hora(fim_bruto),
                "hora_fim": HORA_FIM,
            })

    return linhas, linhas_divergencia, avisos


def gerar_planilha_saida(linhas):
    wb = openpyxl.load_workbook(ARQ_MODELO)
    ws = wb.active

    linha_modelo = 2
    formatos = []
    for col in range(1, 8):
        celula = ws.cell(row=linha_modelo, column=col)
        formatos.append({
            "number_format": celula.number_format,
            "font": copy(celula.font),
            "alignment": copy(celula.alignment),
            "border": copy(celula.border),
            "fill": copy(celula.fill),
        })

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for i, dados in enumerate(linhas):
        r = i + 2
        valores = [
            dados["codigo"],
            dados["preco_site"],
            dados["preco_campanha"],
            dados["data_inicio"],
            dados["hora_inicio"],
            dados["data_fim"],
            dados["hora_fim"],
        ]
        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=r, column=col, value=valor)
            f = formatos[col - 1]
            celula.number_format = f["number_format"]
            celula.font = f["font"]
            celula.alignment = f["alignment"]
            celula.border = f["border"]
            celula.fill = f["fill"]

    wb.save(ARQ_SAIDA)


def gerar_planilha_divergencia(linhas_divergencia):
    wb = openpyxl.load_workbook(ARQ_MODELO_DIVERGENCIA)
    ws = wb.active

    linha_modelo = 2
    formatos = []
    for col in range(1, 3):
        celula = ws.cell(row=linha_modelo, column=col)
        formatos.append({
            "number_format": celula.number_format,
            "font": copy(celula.font),
            "alignment": copy(celula.alignment),
            "border": copy(celula.border),
            "fill": copy(celula.fill),
        })

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for i, dados in enumerate(linhas_divergencia):
        r = i + 2
        valores = [dados["codigo"], dados["preco"]]
        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=r, column=col, value=valor)
            f = formatos[col - 1]
            celula.number_format = f["number_format"]
            celula.font = f["font"]
            celula.alignment = f["alignment"]
            celula.border = f["border"]
            celula.fill = f["fill"]

    wb.save(ARQ_SAIDA_DIVERGENCIA)


def main():
    print("Lendo 'Base dos produtos para alteração.xlsx'...")
    produtos_alterar = ler_produtos_para_alterar()
    print(f"  {len(produtos_alterar)} código(s) PAI para processar.")

    print("Lendo 'PRODUTOS BASE.xlsx' (pode demorar um pouco)...")
    codigos_existentes, filhos_por_pai = ler_produtos_base()

    print("Lendo 'BASE DE PRODUTOS SITE SPORTBAY.xlsx'...")
    preco_site_por_codigo, fim_promocao_por_codigo = ler_site_sportbay()

    linhas, linhas_divergencia, avisos = montar_linhas(
        produtos_alterar, codigos_existentes, filhos_por_pai,
        preco_site_por_codigo, fim_promocao_por_codigo,
    )

    print(f"Gerando '{ARQ_SAIDA.name}' com {len(linhas)} linha(s)...")
    gerar_planilha_saida(linhas)

    if linhas_divergencia:
        print(f"Gerando '{ARQ_SAIDA_DIVERGENCIA.name}' com {len(linhas_divergencia)} linha(s) de divergência...")
        gerar_planilha_divergencia(linhas_divergencia)

    if avisos:
        print(f"\n{len(avisos)} aviso(s) encontrado(s):")
        for aviso in avisos:
            print(f"  - {aviso}")
    else:
        print("\nNenhum aviso. Todos os códigos foram encontrados.")

    print(f"\nConcluído. Arquivo salvo em: {ARQ_SAIDA}")


if __name__ == "__main__":
    main()
