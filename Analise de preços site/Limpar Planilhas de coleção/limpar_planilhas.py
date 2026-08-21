"""
Consolida planilhas de "Produtos Associados" exportadas do painel Marketplace
em UM único arquivo:
  - remove linhas vazias, cabeçalhos repetidos e o bloco de metadados do topo;
  - mantém uma linha por produto (pai), descartando as linhas de variação extras;
  - remove o prefixo "PAI-" dos códigos da coluna C e depois confere cada código
    na base de pais: se ele aparece na coluna B (SKU na Plataforma) e a coluna A
    (Código Auxiliar) for diferente, troca pelo código da coluna A;
  - acrescenta a coluna G "Coleção" com o texto "está na coleção <categoria>",
    lida automaticamente da linha "Categoria: ..." de cada relatório;
  - junta todos os arquivos numa aba só, com o cabeçalho aparecendo uma única vez;
  - preserva a formatação original das células.

USO (no terminal, de qualquer pasta):
    pip install openpyxl
    python limpar_planilhas.py

Sem argumentos ele varre a pasta do próprio script (e a pasta atual, se for outra)
e pega SÓ os arquivos que começam com "produtos_associados" — bases, resultados e
outros .xlsx que estejam por perto são ignorados.

Ou passando arquivos/pastas específicos:
    python limpar_planilhas.py "C:/caminho/arquivo.xlsx" "C:/outro.xlsx"
    python limpar_planilhas.py "C:/caminho/da/pasta"

Gera um único arquivo (padrão: produtos_associados_CONSOLIDADO.xlsx).
Os arquivos originais NUNCA são alterados.
"""

import os
import re
import sys
from copy import copy
from itertools import groupby
from pathlib import Path

import openpyxl

# ------------------------------------------------------------------ ajustes
ABA = "Produtos Associados"     # None = usa sempre a primeira aba de cada arquivo
N_COLS = 6                      # colunas de dados do relatório (A..F)

NOME_SAIDA = "produtos_associados_CONSOLIDADO.xlsx"
UMA_LINHA_POR_PRODUTO = True    # False = mantém todas as variações de cada pai
ADICIONAR_COLUNA_COLECAO = True
TEXTO_COLECAO = "está na coleção {categoria}"
LARGURA_COL_G = 34.875

COL_SKU_PAI = 3                 # coluna C
PREFIXOS_SKU_PAI = ("PAI-",)    # prefixos a remover do "SKU do Pai" (vazio = não mexe)

# Caminhos resolvidos a partir da pasta deste script — hoje
# "Analise de preços site\Limpar Planilhas de coleção".
PASTA_SCRIPT = Path(__file__).resolve().parent
PASTA_RAIZ = PASTA_SCRIPT.parent            # raiz "Analise de preços site"

# Base de pais: coluna A = Código Auxiliar, coluna B = SKU na Plataforma
PASTA_BASE = PASTA_SCRIPT / "Base dos Pais"
ARQUIVO_BASE = None             # None = procura o .xlsx da pasta acima
ABA_BASE = "BASE DE PRODUTOS SITE SPORTBAY"   # None = primeira aba
COL_BASE_AUXILIAR = 1           # coluna A
COL_BASE_PLATAFORMA = 2         # coluna B
IGNORAR = ("_LIMPO", "_CONSOLIDADO")   # arquivos já processados não entram de novo
PREFIXO_ARQUIVO = "produtos_associados"  # ao varrer pastas, só entram estes relatórios
# ---------------------------------------------------------------------------


def carregar_base() -> dict:
    """Devolve {SKU na Plataforma (col. B) -> Código Auxiliar (col. A)}."""
    if ARQUIVO_BASE:
        caminho = Path(ARQUIVO_BASE)
    else:
        pasta = Path(PASTA_BASE)
        if not pasta.is_dir():
            print(f"[aviso] pasta da base não encontrada: {pasta}")
            print("        os códigos ficarão apenas sem o prefixo PAI-")
            return {}
        candidatos = [
            x for x in sorted(pasta.glob("*.xlsx"))
            if not x.name.startswith("~$") and "BASE" in x.stem.upper()
        ] or [x for x in sorted(pasta.glob("*.xlsx")) if not x.name.startswith("~$")]
        if not candidatos:
            print(f"[aviso] nenhum .xlsx na pasta da base: {pasta}")
            return {}
        caminho = candidatos[0]

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[ABA_BASE] if (ABA_BASE and ABA_BASE in wb.sheetnames) else wb.worksheets[0]
    mapa = {}
    for linha in ws.iter_rows(min_row=2, max_col=max(COL_BASE_AUXILIAR, COL_BASE_PLATAFORMA),
                              values_only=True):
        aux = linha[COL_BASE_AUXILIAR - 1]
        plat = linha[COL_BASE_PLATAFORMA - 1]
        if aux in (None, "") or plat in (None, ""):
            continue
        mapa[str(plat).strip()] = str(aux).strip()
    wb.close()
    print(f"Base de pais: {len(mapa)} códigos lidos de {caminho.name}")
    return mapa


def categoria_do_nome(caminho: Path) -> str:
    """Fallback: deduz a categoria pelo nome do arquivo."""
    m = re.search(r"produtos_associados_(.+?)(?:_\d{10,})?(?:_LIMPO)?$", caminho.stem)
    return m.group(1).replace("_", " ").strip() if m else "sem categoria"


def limpar_planilha(caminho: Path):
    """Devolve (worksheet limpa, categoria). Trabalha só em memória."""
    wb = openpyxl.load_workbook(caminho)
    ws = wb[ABA] if (ABA and ABA in wb.sheetnames) else wb.worksheets[0]

    categoria = None
    apagar = []
    cabecalho_mantido = False

    for r in range(1, ws.max_row + 1):
        s = [
            ("" if ws.cell(r, c).value is None else str(ws.cell(r, c).value)).strip()
            for c in range(1, N_COLS + 1)
        ]
        if all(x == "" for x in s):
            apagar.append(r)                       # linha totalmente vazia
        elif s[0] == "Merchant" and s[1] == "ShortHash do Pai":
            if cabecalho_mantido:
                apagar.append(r)                   # cabeçalho repetido
            else:
                cabecalho_mantido = True           # mantém apenas o primeiro
        elif all(x == "" for x in s[1:]):
            m = re.match(r"^Categoria:\s*(.+)$", s[0])
            if m:
                categoria = m.group(1).strip()
            apagar.append(r)                       # metadados do relatório
        elif UMA_LINHA_POR_PRODUTO and s[0] == "":
            apagar.append(r)                       # variação extra do mesmo pai

    # apaga de baixo para cima, em blocos contíguos
    blocos = []
    for _, g in groupby(enumerate(apagar), lambda x: x[1] - x[0]):
        g = [v for _, v in g]
        blocos.append((g[0], len(g)))
    for inicio, qtd in reversed(blocos):
        ws.delete_rows(inicio, qtd)

    return ws, (categoria or categoria_do_nome(caminho))


def ajustar_sku_pai(valor, coluna, base, contadores):
    """Tira o prefixo 'PAI-' e, pela base, troca o SKU da plataforma pelo auxiliar."""
    if coluna != COL_SKU_PAI or not isinstance(valor, str):
        return valor

    cod = valor.strip()
    for pref in PREFIXOS_SKU_PAI:
        if cod.upper().startswith(pref.upper()):
            cod = cod[len(pref):]
            contadores["prefixo"] += 1
            break

    if base:
        auxiliar = base.get(cod)
        if auxiliar is None:
            contadores["fora_da_base"].add(cod)
        elif auxiliar != cod:
            contadores["trocado"] += 1
            return auxiliar
    return cod


def fonte_padrao(ws):
    """Fonte "Normal" do arquivo de origem (o painel não estiliza célula a célula)."""
    try:
        return copy(ws.parent._named_styles["Normal"].font)
    except Exception:
        return None


def copiar_estilo(origem, destino, padrao=None):
    """Copia o visual da célula entre workbooks diferentes.

    Não dá para usar `destino._style = copy(origem._style)`: o StyleArray guarda
    apenas ÍNDICES para as tabelas de estilo do arquivo de origem, que não existem
    no workbook novo. O save estoura com IndexError e grava um .xlsx corrompido.
    """
    if not origem.has_style:
        if padrao is not None:             # sem estilo próprio: só a fonte padrão
            destino.font = copy(padrao)
        return
    destino.font = copy(origem.font)
    destino.fill = copy(origem.fill)
    destino.border = copy(origem.border)
    destino.alignment = copy(origem.alignment)
    destino.protection = copy(origem.protection)
    destino.number_format = origem.number_format


def copiar_linha(origem_ws, r_origem, destino_ws, r_destino, n_cols,
                 base=None, contadores=None, cabecalho=False, padrao=None):
    for c in range(1, n_cols + 1):
        o = origem_ws.cell(r_origem, c)
        d = destino_ws.cell(r_destino, c)
        d.value = o.value if cabecalho else ajustar_sku_pai(o.value, c, base, contadores)
        copiar_estilo(o, d, padrao)        # fonte, preenchimento, borda, formato
    return destino_ws.cell(r_destino, n_cols)   # última célula, serve de modelo


def consolidar(arquivos, pasta_saida: Path) -> Path:
    base = carregar_base()
    contadores = {"prefixo": 0, "trocado": 0, "fora_da_base": set()}

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ABA or "Produtos Associados"

    linha = 0
    total_por_colecao = []

    for caminho in arquivos:
        ws, categoria = limpar_planilha(caminho)
        if ws.max_row < 2:
            print(f"[aviso] {caminho.name}: sem linhas de dados, ignorado")
            continue

        padrao = fonte_padrao(ws)
        if linha == 0:                                   # cabeçalho, uma vez só
            linha += 1
            modelo = copiar_linha(ws, 1, ws_out, linha, N_COLS,
                                  cabecalho=True, padrao=padrao)
            if ADICIONAR_COLUNA_COLECAO:
                g = ws_out.cell(linha, N_COLS + 1)
                g.value = "Coleção"
                g._style = copy(modelo._style)
            for letra, dim in ws.column_dimensions.items():
                if dim.width:
                    ws_out.column_dimensions[letra].width = dim.width
            if ADICIONAR_COLUNA_COLECAO:
                ws_out.column_dimensions["G"].width = LARGURA_COL_G

        texto = TEXTO_COLECAO.format(categoria=categoria)
        for r in range(2, ws.max_row + 1):
            linha += 1
            modelo = copiar_linha(ws, r, ws_out, linha, N_COLS, base, contadores,
                                  padrao=padrao)
            if ADICIONAR_COLUNA_COLECAO:
                g = ws_out.cell(linha, N_COLS + 1)
                g.value = texto
                g._style = copy(modelo._style)

        total_por_colecao.append((categoria, ws.max_row - 1, caminho.name))
        print(f"{caminho.name}: {ws.max_row - 1} linhas | {categoria}")

    destino = pasta_saida / NOME_SAIDA
    parcial = destino.with_suffix(".xlsx.tmp")   # se o save falhar, não deixa
    wb_out.save(parcial)                         # um arquivo corrompido no lugar
    os.replace(parcial, destino)
    print(f"\n{len(total_por_colecao)} arquivo(s) unidos -> {destino}")
    print(f"Total: {max(linha - 1, 0)} linhas de produto + cabeçalho")
    print(f'Coluna C: {contadores["prefixo"]} com prefixo "PAI-" removido, '
          f'{contadores["trocado"]} trocados pelo Código Auxiliar')
    fora = sorted(contadores["fora_da_base"])
    if fora:
        print(f"[aviso] {len(fora)} código(s) não encontrado(s) na base "
              f"(mantidos como estavam): {', '.join(fora[:20])}"
              + (" ..." if len(fora) > 20 else ""))
    return destino


def e_relatorio(p: Path) -> bool:
    """Só os relatórios exportados do painel entram na consolidação."""
    return p.name.lower().startswith(PREFIXO_ARQUIVO.lower())


def pastas_padrao() -> list:
    """Sem argumentos: procura na pasta do script e, se for outra, na pasta atual."""
    pastas = [Path(__file__).resolve().parent]
    atual = Path.cwd().resolve()
    if atual != pastas[0]:
        pastas.append(atual)
    return pastas


def coletar(args) -> list:
    alvos = []
    for a in (args or pastas_padrao()):
        p = Path(a)
        if p.is_dir():
            # numa pasta, só os "produtos_associados*.xlsx" — evita varrer
            # bases e resultados que estejam na mesma pasta
            alvos += [x for x in sorted(p.glob("*.xlsx")) if e_relatorio(x)]
        elif p.is_file():
            alvos.append(p)              # arquivo passado na mão: respeita a escolha
        else:
            print(f"[aviso] não encontrado: {a}")

    vistos, saida = set(), []
    for p in alvos:
        if any(t in p.stem for t in IGNORAR) or p.name.startswith("~$"):
            continue
        chave = p.resolve()
        if chave in vistos:
            continue
        vistos.add(chave)
        saida.append(p)
    return saida


if __name__ == "__main__":
    arquivos = coletar(sys.argv[1:])
    if not arquivos:
        print(f'Nenhum "{PREFIXO_ARQUIVO}*.xlsx" encontrado para processar.')
        sys.exit(0)
    print(f"{len(arquivos)} arquivo(s) para processar: "
          + ", ".join(p.name for p in arquivos))
    pasta = Path(arquivos[0]).parent
    try:
        consolidar(arquivos, pasta)
    except Exception as e:
        print(f"[erro] {e}")
