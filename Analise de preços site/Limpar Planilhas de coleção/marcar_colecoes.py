"""
Cruza a planilha de análise de preços com o consolidado de coleções.

O que faz:
  - lê os códigos da coluna A da "Resultado_Menor_Preco_Sportbay";
  - procura cada um na coluna C ("SKU do Pai") do "produtos_associados_CONSOLIDADO";
  - copia o texto da coluna G ("Coleção") para a coluna J da planilha de preços;
  - se o mesmo código estiver em mais de uma coleção, junta os textos com " | ";
  - salva uma cópia nova preservando toda a formatação, fórmulas e as demais abas.

USO (no terminal):
    pip install openpyxl
    python marcar_colecoes.py

Ou apontando os arquivos na mão:
    python marcar_colecoes.py "C:/.../Resultado_Menor_Preco_Sportbay.xlsx" "C:/.../produtos_associados_CONSOLIDADO.xlsx"

Os arquivos originais NUNCA são alterados.
"""

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# ------------------------------------------------------------------ ajustes
PASTA_PRECOS = r"C:\Users\Protork\Desktop\Analise de preços site"
ARQ_PRECOS = None               # None = procura "Resultado_Menor_Preco*.xlsx" na pasta acima
ABA_PRECOS = "Resultado"        # None = primeira aba
COL_PRECOS_CODIGO = 1           # coluna A - Código Fabrica
COL_PRECOS_DESTINO = 10         # coluna J - onde entra o texto da coleção

PASTA_CONSOLIDADO = PASTA_PRECOS   # troque se o consolidado ficar em outra pasta
ARQ_CONSOLIDADO = None          # None = procura "*CONSOLIDADO*.xlsx" na pasta acima
ABA_CONSOLIDADO = "Produtos Associados"
COL_CONS_CODIGO = 3             # coluna C - SKU do Pai
COL_CONS_COLECAO = 7            # coluna G - Coleção

SEPARADOR = " | "               # usado quando o código está em mais de uma coleção
TEXTO_SEM_COLECAO = ""          # ex.: "não está em nenhuma coleção" se quiser preencher
SUFIXO_SAIDA = "_COM_COLECOES"
# ---------------------------------------------------------------------------


def achar_arquivo(pasta: str, padroes, descricao: str) -> Path:
    p = Path(pasta)
    if not p.is_dir():
        raise SystemExit(f"[erro] pasta não encontrada: {p}")
    for padrao in padroes:
        achados = [x for x in sorted(p.glob(padrao)) if not x.name.startswith("~$")]
        if achados:
            return achados[0]
    raise SystemExit(f"[erro] não achei a planilha de {descricao} em {p}")


def normalizar(valor) -> str:
    """Deixa o código comparável: sem espaços, sem '.0' de número, em maiúsculas."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).strip().upper()


def carregar_colecoes(caminho: Path) -> dict:
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[ABA_CONSOLIDADO] if (ABA_CONSOLIDADO and ABA_CONSOLIDADO in wb.sheetnames) else wb.worksheets[0]

    mapa = defaultdict(list)
    ncols = max(COL_CONS_CODIGO, COL_CONS_COLECAO)
    for linha in ws.iter_rows(min_row=2, max_col=ncols, values_only=True):
        cod = normalizar(linha[COL_CONS_CODIGO - 1])
        colecao = linha[COL_CONS_COLECAO - 1]
        if not cod or not colecao:
            continue
        texto = str(colecao).strip()
        if texto not in mapa[cod]:                 # evita repetir a mesma coleção
            mapa[cod].append(texto)
    wb.close()
    print(f"Consolidado: {len(mapa)} códigos lidos de {caminho.name}")
    return mapa


def marcar(caminho_precos: Path, mapa: dict) -> Path:
    wb = openpyxl.load_workbook(caminho_precos)   # mantém formatação e todas as abas
    ws = wb[ABA_PRECOS] if (ABA_PRECOS and ABA_PRECOS in wb.sheetnames) else wb.worksheets[0]

    achados = 0
    multiplas = []
    sem_colecao = []

    for r in range(2, ws.max_row + 1):
        cod = normalizar(ws.cell(r, COL_PRECOS_CODIGO).value)
        if not cod:
            continue
        colecoes = mapa.get(cod)
        if colecoes:
            ws.cell(r, COL_PRECOS_DESTINO).value = SEPARADOR.join(colecoes)
            achados += 1
            if len(colecoes) > 1:
                multiplas.append(cod)
        else:
            sem_colecao.append(cod)
            if TEXTO_SEM_COLECAO:
                ws.cell(r, COL_PRECOS_DESTINO).value = TEXTO_SEM_COLECAO

    destino = caminho_precos.with_name(caminho_precos.stem + SUFIXO_SAIDA + ".xlsx")
    wb.save(destino)

    print(f"{caminho_precos.name}: {achados} código(s) marcado(s), "
          f"{len(sem_colecao)} sem coleção")
    if multiplas:
        print(f"[info] {len(multiplas)} em mais de uma coleção: {', '.join(multiplas[:10])}"
              + (" ..." if len(multiplas) > 10 else ""))
    print(f"=> {destino}")
    return destino


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        precos, consolidado = Path(sys.argv[1]), Path(sys.argv[2])
    else:
        precos = Path(ARQ_PRECOS) if ARQ_PRECOS else achar_arquivo(
            PASTA_PRECOS, ["Resultado_Menor_Preco*.xlsx", "Resultado*.xlsx"], "preços")
        consolidado = Path(ARQ_CONSOLIDADO) if ARQ_CONSOLIDADO else achar_arquivo(
            PASTA_CONSOLIDADO, ["*CONSOLIDADO*.xlsx", "produtos_associados*.xlsx"], "coleções")

    print(f"Preços:      {precos}")
    print(f"Consolidado: {consolidado}\n")
    marcar(precos, carregar_colecoes(consolidado))
