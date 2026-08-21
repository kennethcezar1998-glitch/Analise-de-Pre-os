# -*- coding: utf-8 -*-
"""
Bot de limpeza da planilha "BI".

Le o export bruto "BI.xlsx" (layout escuro, com celulas mescladas e linha de
rodape "Total"), cruza cada variacao com a "PRODUTOS BASE.xlsx" para descobrir o
"Produto PAI" e gera "BI DD-MM-AAAA.xlsx" no mesmo formato de "BI - Limpo.xlsx":

  Sheet1    -> todas as variacoes, com Pai / Total Variacao / Total Real,
               ordenadas por "Total Variacao" (maior -> menor)
  Planilha1 -> apenas a variacao mais vendida de cada Pai,
               ordenada por "Total Real" (maior -> menor)

Uso:
    python limpar_bi.py
    python limpar_bi.py --bi "BI.xlsx" --base "...\\PRODUTOS BASE.xlsx" --out "."
"""

from __future__ import annotations

import argparse
import os
import pickle
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Caminhos padrao
# --------------------------------------------------------------------------- #
AQUI = os.path.dirname(os.path.abspath(__file__))
BI_PADRAO = os.path.join(AQUI, "BI.xlsx")
BASE_PADRAO = os.path.abspath(
    os.path.join(
        AQUI, "..", "..",
        "Planilhas com as Bases de Produtos - Site e ERP",
        "PRODUTOS BASE.xlsx",
    )
)
CACHE_BASE = os.path.join(AQUI, ".cache_produtos_base.pkl")

# --------------------------------------------------------------------------- #
# Formatacao (copiada de "BI - Limpo.xlsx")
# --------------------------------------------------------------------------- #
CABECALHOS = [
    "Pai", "Código Fabrica", "Código Aux",
    "Produto", "Categoria", "Total Variação", "Total Real",
]
LARGURAS = {
    "A": 12.140625, "B": 16.7109375, "C": 19.42578125,
    "D": 107.140625, "E": 34.140625, "F": 24.85546875, "G": 24.85546875,
}
COLS_CENTRALIZADAS = {1, 5, 6, 7}      # Pai, Categoria, Total Variacao, Total Real

FONTE_CABECALHO = Font(name="Calibri", size=11, color="FF00FF00")
FUNDO_CABECALHO = PatternFill(fill_type="solid", start_color="FF000000",
                              end_color="FF000000")
FONTE_DADOS = Font(name="Calibri", size=11)
BORDA = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))
CENTRO = Alignment(horizontal="center")


# --------------------------------------------------------------------------- #
# Utilitarios
# --------------------------------------------------------------------------- #
def normalizar(valor):
    """minusculas, sem acento, sem espacos duplicados - para comparar rotulos."""
    if valor is None:
        return ""
    t = unicodedata.normalize("NFKD", str(valor))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return " ".join(t.lower().split())


def texto(valor):
    """Converte celula em texto preservando codigos (sem virar 1.2345e+11)."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


PADRAO_MILHAR = re.compile(r"^-?\d{1,3}(\.\d{3})+$")


def numero(valor):
    """Le o 'Total' do BI: '1', '143', '1.052', '80.899' -> int."""
    if valor is None or valor == "":
        return 0
    if isinstance(valor, (int, float)):
        return int(round(valor))
    t = str(valor).strip().replace("\xa0", "").replace(" ", "")
    if PADRAO_MILHAR.match(t):
        t = t.replace(".", "")
    elif "," in t:
        t = t.replace(".", "").replace(",", ".")
    try:
        return int(round(float(t)))
    except ValueError:
        return 0


# --------------------------------------------------------------------------- #
# Leitura do BI bruto
# --------------------------------------------------------------------------- #
PADRAO_TOTAL = re.compile(r"^-?\d{1,3}(\.\d{3})*$")
PAPEIS = {
    "codigo fabrica": "fab",
    "codigo fabricante": "fab",
    "codigo aux": "aux",
    "codigo auxiliar": "aux",
    "produto": "prod",
    "categoria": "cat",
    "total": "total",
}


def _mapear_colunas(grade, linhas_dados):
    """Descobre em quais colunas estao Fabrica / Aux / Produto / Categoria / Total.

    O export do BI poe o rotulo numa coluna e o dado em outra (celulas
    mescladas), entao cada coluna de dado recebe o rotulo mais proximo.
    """
    amostra = linhas_dados[:200]
    freq = defaultdict(int)
    for r in amostra:
        for c, v in enumerate(grade[r], 1):
            if texto(v):
                freq[c] += 1
    limite = max(1, len(amostra) * 0.2)
    colunas = sorted(c for c, n in freq.items() if n >= limite)

    # rotulos: tudo que estiver acima da primeira linha de dados
    rotulos = {}
    for r in range(linhas_dados[0]):
        for c, v in enumerate(grade[r], 1):
            papel = PAPEIS.get(normalizar(v))
            if papel and c not in rotulos:
                rotulos[c] = papel

    mapa = {}
    if rotulos:
        for c in colunas:
            perto = min(rotulos, key=lambda rc: (abs(rc - c), rc))
            if abs(perto - c) <= 2 and rotulos[perto] not in mapa:
                mapa[rotulos[perto]] = c

    if set(mapa) != {"fab", "aux", "prod", "cat", "total"}:
        if len(colunas) != 5:
            raise SystemExit(
                "Nao consegui identificar as colunas do BI (achei %d colunas de "
                "dados: %s). Confira o layout do arquivo." % (len(colunas), colunas)
            )
        mapa = dict(zip(("fab", "aux", "prod", "cat", "total"), colunas))
    return mapa


def ler_bi(caminho):
    """Devolve [(codigo_fabrica, codigo_aux, produto, categoria, total, ordem)]."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    grade = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    linhas_dados = []
    for i, row in enumerate(grade):
        cheias = [texto(v) for v in row if texto(v)]
        if len(cheias) < 3:
            continue                                # cabecalho / rodape "Total"
        if normalizar(cheias[0]) == "total":
            continue
        if not PADRAO_TOTAL.match(cheias[-1]):
            continue                                # linha sem quantidade
        linhas_dados.append(i)

    if not linhas_dados:
        raise SystemExit("Nenhuma linha de dados encontrada em %s" % caminho)

    m = _mapear_colunas(grade, linhas_dados)
    registros = []
    for ordem, i in enumerate(linhas_dados):
        row = grade[i]

        def pega(c):
            return row[c - 1] if c - 1 < len(row) else None

        registros.append((
            texto(pega(m["fab"])),
            texto(pega(m["aux"])),
            texto(pega(m["prod"])),
            texto(pega(m["cat"])),
            numero(pega(m["total"])),
            ordem,
        ))
    return registros


# --------------------------------------------------------------------------- #
# Leitura da PRODUTOS BASE (Codigo Auxiliar -> Produto PAI)
# --------------------------------------------------------------------------- #
def ler_base(caminho, usar_cache=True):
    assinatura = (os.path.getmtime(caminho), os.path.getsize(caminho))
    if usar_cache and os.path.exists(CACHE_BASE):
        try:
            with open(CACHE_BASE, "rb") as fh:
                gravado, mapa = pickle.load(fh)
            if gravado == assinatura:
                print("  (cache) %d codigos carregados" % len(mapa))
                return mapa
        except Exception:
            pass

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas)

    col_aux = col_pai = None
    for c, v in enumerate(cabecalho):
        n = normalizar(v)
        if n == "codigo auxiliar" and col_aux is None:
            col_aux = c
        elif n == "produto pai" and col_pai is None:
            col_pai = c
    if col_aux is None:
        col_aux = 1                                 # coluna B
    if col_pai is None:
        col_pai = 3                                 # coluna D

    mapa = {}
    for row in linhas:
        if col_aux >= len(row):
            continue
        chave = texto(row[col_aux])
        if not chave or chave in mapa:
            continue
        mapa[chave] = texto(row[col_pai]) if col_pai < len(row) else ""
    wb.close()

    try:
        with open(CACHE_BASE, "wb") as fh:
            pickle.dump((assinatura, mapa), fh)
    except Exception:
        pass
    return mapa


# --------------------------------------------------------------------------- #
# Regras de negocio
# --------------------------------------------------------------------------- #
def montar(registros, mapa_pai):
    sem_pai = []
    linhas = []
    for fab, aux, prod, cat, total, ordem in registros:
        pai = mapa_pai.get(aux, "")
        if not pai:
            sem_pai.append((fab, aux, prod))
        # sem Pai a variacao vira grupo dela mesma (nao some da planilha)
        grupo = pai or ("\x00" + aux)
        linhas.append({
            "pai": pai, "fab": fab, "aux": aux, "prod": prod, "cat": cat,
            "tv": total, "grupo": grupo, "ordem": ordem,
        })

    total_real = defaultdict(int)
    for l in linhas:
        total_real[l["grupo"]] += l["tv"]
    for l in linhas:
        l["tr"] = total_real[l["grupo"]]

    # Sheet1: todas as variacoes, "Total Variacao" decrescente
    # (empate mantem a ordem original do BI)
    sheet1 = sorted(linhas, key=lambda l: (-l["tv"], l["ordem"]))

    # Planilha1: a variacao mais vendida de cada Pai, "Total Real" decrescente
    melhor = {}
    for l in sheet1:
        melhor.setdefault(l["grupo"], l)
    planilha1 = sorted(melhor.values(), key=lambda l: (-l["tr"], l["ordem"]))

    return sheet1, planilha1, sem_pai


# --------------------------------------------------------------------------- #
# Escrita
# --------------------------------------------------------------------------- #
def escrever_aba(ws, linhas):
    ws.sheet_view.showGridLines = False

    for c, titulo in enumerate(CABECALHOS, 1):
        cel = ws.cell(1, c, titulo)
        cel.font = FONTE_CABECALHO
        cel.fill = FUNDO_CABECALHO
        cel.alignment = CENTRO
        cel.border = BORDA

    for i, l in enumerate(linhas, 2):
        valores = (l["pai"], l["fab"], l["aux"], l["prod"], l["cat"],
                   l["tv"], l["tr"])
        for c, valor in enumerate(valores, 1):
            cel = ws.cell(i, c, valor)
            cel.font = FONTE_DADOS
            cel.border = BORDA
            if c in COLS_CENTRALIZADAS:
                cel.alignment = CENTRO
            if c == 3:                              # Codigo Aux fica como texto
                cel.number_format = "@"

    for letra, largura in LARGURAS.items():
        ws.column_dimensions[letra].width = largura

    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(CABECALHOS)),
                                      max(1, len(linhas) + 1))


def gravar(sheet1, planilha1, destino):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    escrever_aba(ws1, sheet1)
    escrever_aba(wb.create_sheet("Planilha1"), planilha1)
    wb.save(destino)


# --------------------------------------------------------------------------- #
def main():
    p = argparse.ArgumentParser(description="Limpa a planilha BI.")
    p.add_argument("--bi", default=BI_PADRAO, help="BI.xlsx bruto")
    p.add_argument("--base", default=BASE_PADRAO, help="PRODUTOS BASE.xlsx")
    p.add_argument("--out", default=AQUI, help="pasta de saida")
    p.add_argument("--data", default=None,
                   help="data usada no nome do arquivo (DD-MM-AAAA)")
    p.add_argument("--sem-cache", action="store_true",
                   help="ignora o cache da PRODUTOS BASE")
    a = p.parse_args()

    for caminho, rotulo in ((a.bi, "BI"), (a.base, "PRODUTOS BASE")):
        if not os.path.exists(caminho):
            raise SystemExit("Arquivo nao encontrado (%s): %s" % (rotulo, caminho))

    print("Lendo %s ..." % os.path.basename(a.bi))
    registros = ler_bi(a.bi)
    print("  %d variacoes | soma dos totais: %d"
          % (len(registros), sum(r[4] for r in registros)))

    print("Lendo %s ..." % os.path.basename(a.base))
    mapa_pai = ler_base(a.base, usar_cache=not a.sem_cache)
    print("  %d codigos auxiliares na base" % len(mapa_pai))

    sheet1, planilha1, sem_pai = montar(registros, mapa_pai)
    print("Sheet1: %d linhas | Planilha1: %d pais"
          % (len(sheet1), len(planilha1)))

    dia = a.data or date.today().strftime("%d-%m-%Y")
    destino = os.path.join(a.out, "BI %s.xlsx" % dia)
    gravar(sheet1, planilha1, destino)
    print("Gerado: %s" % destino)

    if sem_pai:
        log = os.path.join(a.out, "BI %s - sem pai.txt" % dia)
        with open(log, "w", encoding="utf-8") as fh:
            fh.write("Codigos sem 'Produto PAI' na PRODUTOS BASE "
                     "(mantidos na planilha, agrupados sozinhos)\n\n")
            for fab, aux, prod in sem_pai:
                fh.write("%s\t%s\t%s\n" % (fab, aux, prod))
        print("ATENCAO: %d variacao(oes) sem Pai na base -> %s"
              % (len(sem_pai), log))


if __name__ == "__main__":
    sys.exit(main())
