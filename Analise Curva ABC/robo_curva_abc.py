"""
Robo de Analise de Curva ABC - Site Sportbay

Le a planilha "Analise Curvas ABC.xlsx" (colunas: Pai, Codigo Fabrica, Codigo Aux,
Produto, Categoria, Total, Valor Total (NF)) e gera "Curva_ABC_Sportbay.xlsx" com:
  - Classificacao ABC por Valor (receita) e por Quantidade vendida
  - Resumo consolidado por classe (A/B/C)
  - Resumo por categoria de produto
  - Graficos de Pareto (top 30 produtos) para Valor e Quantidade

Basta rodar novamente este script sempre que a planilha de origem for atualizada.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

PASTA = Path(__file__).parent
ARQUIVO_ENTRADA = PASTA / "Analise Curvas ABC.xlsx"
ARQUIVO_SAIDA = PASTA / "Curva_ABC_Sportbay.xlsx"

LIMITE_A = 0.80
LIMITE_B = 0.95

FONTE = "Arial"
COR_A = "C6EFCE"  # verde
COR_B = "FFEB9C"  # amarelo
COR_C = "FFC7CE"  # vermelho
COR_HEADER = "1F4E78"


def carregar_dados() -> pd.DataFrame:
    if not ARQUIVO_ENTRADA.exists():
        sys.exit(f"Arquivo nao encontrado: {ARQUIVO_ENTRADA}")

    df = pd.read_excel(ARQUIVO_ENTRADA, sheet_name="Plan1", engine="openpyxl")

    df.columns = [str(c).strip() for c in df.columns]
    df = df.rename(columns={"Valor Total (NF)": "Valor_Total", "Total": "Quantidade"})

    linhas_antes = len(df)
    df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce")
    df["Valor_Total"] = pd.to_numeric(df["Valor_Total"], errors="coerce")
    df = df.dropna(subset=["Pai", "Quantidade", "Valor_Total"])
    df = df[df["Pai"].astype(str).str.strip() != "#N/A"]
    descartadas = linhas_antes - len(df)
    if descartadas:
        print(f"Aviso: {descartadas} linha(s) invalida(s)/incompleta(s) descartadas.")

    df["Categoria"] = df["Categoria"].fillna("SEM CATEGORIA")
    df.loc[df["Categoria"].astype(str).str.strip().isin(["", "#N/A"]), "Categoria"] = "SEM CATEGORIA"

    return df


def classificar_abc(df: pd.DataFrame, coluna: str, sufixo: str) -> pd.DataFrame:
    """Ordena por `coluna` desc, calcula % acumulado e classe A/B/C."""
    tmp = df.sort_values(coluna, ascending=False).copy()
    total = tmp[coluna].sum()
    tmp[f"Pct_{sufixo}"] = tmp[coluna] / total
    tmp[f"PctAcum_{sufixo}"] = tmp[coluna].cumsum() / total
    tmp[f"Classe_{sufixo}"] = pd.cut(
        tmp[f"PctAcum_{sufixo}"], [0, LIMITE_A, LIMITE_B, 1.0000001],
        labels=["A", "B", "C"],
    )
    return tmp[["Pai", f"Pct_{sufixo}", f"PctAcum_{sufixo}", f"Classe_{sufixo}"]]


def montar_tabela(df: pd.DataFrame) -> pd.DataFrame:
    abc_valor = classificar_abc(df, "Valor_Total", "Valor")
    abc_qtd = classificar_abc(df, "Quantidade", "Qtd")

    resultado = (
        df.merge(abc_valor, on="Pai", how="left")
          .merge(abc_qtd, on="Pai", how="left")
          .sort_values("Valor_Total", ascending=False)
          .reset_index(drop=True)
    )
    return resultado


def montar_resumo(resultado: pd.DataFrame) -> pd.DataFrame:
    linhas = []
    for base, col_valor, col_classe in [
        ("Valor (Receita)", "Valor_Total", "Classe_Valor"),
        ("Quantidade", "Quantidade", "Classe_Qtd"),
    ]:
        total_geral = resultado[col_valor].sum()
        total_skus = len(resultado)
        for classe in ["A", "B", "C"]:
            sub = resultado[resultado[col_classe] == classe]
            linhas.append({
                "Base": base,
                "Classe": classe,
                "Qtd_SKUs": len(sub),
                "Pct_SKUs": len(sub) / total_skus if total_skus else 0,
                "Total_Metrica": sub[col_valor].sum(),
                "Pct_Metrica": sub[col_valor].sum() / total_geral if total_geral else 0,
            })
    return pd.DataFrame(linhas)


def montar_resumo_categoria(resultado: pd.DataFrame) -> pd.DataFrame:
    tab = (
        resultado.groupby(["Categoria", "Classe_Valor"], observed=True)
        .agg(Qtd_SKUs=("Pai", "count"), Valor_Total=("Valor_Total", "sum"))
        .reset_index()
    )
    pivot_qtd = tab.pivot(index="Categoria", columns="Classe_Valor", values="Qtd_SKUs").fillna(0)
    pivot_valor = tab.pivot(index="Categoria", columns="Classe_Valor", values="Valor_Total").fillna(0)
    pivot_qtd.columns = [f"SKUs_{c}" for c in pivot_qtd.columns]
    pivot_valor.columns = [f"Valor_{c}" for c in pivot_valor.columns]
    resumo = pivot_qtd.join(pivot_valor)
    resumo["Total_SKUs"] = resumo[[c for c in resumo.columns if c.startswith("SKUs_")]].sum(axis=1)
    resumo["Total_Valor"] = resumo[[c for c in resumo.columns if c.startswith("Valor_")]].sum(axis=1)
    resumo = resumo.sort_values("Total_Valor", ascending=False).reset_index()
    return resumo


# ---------------------------------------------------------------------------
# Escrita / formatacao do Excel
# ---------------------------------------------------------------------------

def estilizar_cabecalho(ws, ncols, linha=1):
    fill = PatternFill("solid", fgColor=COR_HEADER)
    fonte = Font(name=FONTE, bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        cel = ws.cell(row=linha, column=col)
        cel.fill = fill
        cel.font = fonte
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)


def colorir_classe(ws, col_letra, linha_ini, linha_fim):
    cores = {"A": COR_A, "B": COR_B, "C": COR_C}
    for r in range(linha_ini, linha_fim + 1):
        cel = ws[f"{col_letra}{r}"]
        if cel.value in cores:
            cel.fill = PatternFill("solid", fgColor=cores[cel.value])
            cel.alignment = Alignment(horizontal="center")


def autoajustar_colunas(ws, largura_max=45):
    for col_cells in ws.columns:
        comprimento = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        letra = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letra].width = min(comprimento + 2, largura_max)


def escrever_dataframe(ws, df: pd.DataFrame, linha_ini=1):
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=linha_ini, column=j, value=str(col))
    for i, row in enumerate(df.itertuples(index=False), start=linha_ini + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    for r in range(linha_ini, linha_ini + len(df) + 1):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).font = Font(name=FONTE, size=10)


def aba_curva_abc(wb, resultado):
    ws = wb.create_sheet("Curva ABC")
    colunas = [
        "Pai", "Código Fabrica", "Código Aux", "Produto", "Categoria",
        "Quantidade", "Valor_Total",
        "Pct_Valor", "PctAcum_Valor", "Classe_Valor",
        "Pct_Qtd", "PctAcum_Qtd", "Classe_Qtd",
    ]
    df = resultado[[c for c in colunas if c in resultado.columns]]
    escrever_dataframe(ws, df)

    n = len(df)
    idx = {c: i + 1 for i, c in enumerate(df.columns)}

    for r in range(2, n + 2):
        ws.cell(row=r, column=idx["Valor_Total"]).number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=idx["Quantidade"]).number_format = '#,##0'
        ws.cell(row=r, column=idx["Pct_Valor"]).number_format = '0.00%'
        ws.cell(row=r, column=idx["PctAcum_Valor"]).number_format = '0.00%'
        ws.cell(row=r, column=idx["Pct_Qtd"]).number_format = '0.00%'
        ws.cell(row=r, column=idx["PctAcum_Qtd"]).number_format = '0.00%'

    estilizar_cabecalho(ws, len(df.columns))
    colorir_classe(ws, get_column_letter(idx["Classe_Valor"]), 2, n + 1)
    colorir_classe(ws, get_column_letter(idx["Classe_Qtd"]), 2, n + 1)
    autoajustar_colunas(ws)

    tabela = Table(displayName="TabCurvaABC", ref=f"A1:{get_column_letter(len(df.columns))}{n + 1}")
    tabela.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tabela)
    return ws


def aba_resumo(wb, resumo):
    ws = wb.create_sheet("Resumo", 0)
    escrever_dataframe(ws, resumo)
    n = len(resumo)
    idx = {c: i + 1 for i, c in enumerate(resumo.columns)}
    for r in range(2, n + 2):
        ws.cell(row=r, column=idx["Pct_SKUs"]).number_format = '0.0%'
        ws.cell(row=r, column=idx["Pct_Metrica"]).number_format = '0.0%'
        ws.cell(row=r, column=idx["Total_Metrica"]).number_format = '#,##0.00'
    estilizar_cabecalho(ws, len(resumo.columns))
    colorir_classe(ws, get_column_letter(idx["Classe"]), 2, n + 1)
    autoajustar_colunas(ws)
    return ws


def aba_categoria(wb, resumo_cat):
    ws = wb.create_sheet("Resumo por Categoria")
    escrever_dataframe(ws, resumo_cat)
    n = len(resumo_cat)
    idx = {c: i + 1 for i, c in enumerate(resumo_cat.columns)}
    for r in range(2, n + 2):
        for col_nome in resumo_cat.columns:
            if col_nome.startswith("Valor_") or col_nome == "Total_Valor":
                ws.cell(row=r, column=idx[col_nome]).number_format = 'R$ #,##0.00'
    estilizar_cabecalho(ws, len(resumo_cat.columns))
    autoajustar_colunas(ws)
    return ws


def aba_graficos(wb, resultado):
    ws = wb.create_sheet("Graficos")

    top_valor = resultado.sort_values("Valor_Total", ascending=False).head(30).reset_index(drop=True)
    top_qtd = resultado.sort_values("Quantidade", ascending=False).head(30).reset_index(drop=True)

    ws["A1"] = "Top 30 - Pareto por Valor (Receita)"
    ws["A1"].font = Font(name=FONTE, bold=True, size=12)
    ws["A2"] = "Produto"
    ws["B2"] = "Valor_Total"
    ws["C2"] = "PctAcum_Valor"
    for i, row in top_valor.iterrows():
        ws.cell(row=3 + i, column=1, value=str(row["Produto"])[:40])
        ws.cell(row=3 + i, column=2, value=row["Valor_Total"])
        ws.cell(row=3 + i, column=3, value=row["PctAcum_Valor"])

    ws["F1"] = "Top 30 - Pareto por Quantidade"
    ws["F1"].font = Font(name=FONTE, bold=True, size=12)
    ws["F2"] = "Produto"
    ws["G2"] = "Quantidade"
    ws["H2"] = "PctAcum_Qtd"
    for i, row in top_qtd.iterrows():
        ws.cell(row=3 + i, column=6, value=str(row["Produto"])[:40])
        ws.cell(row=3 + i, column=7, value=row["Quantidade"])
        ws.cell(row=3 + i, column=8, value=row["PctAcum_Qtd"])

    n = len(top_valor)

    bar = BarChart()
    bar.title = "Pareto - Valor (Receita) - Top 30 SKUs"
    bar.y_axis.title = "Valor Total (R$)"
    bar.x_axis.title = "Produto"
    dados = Reference(ws, min_col=2, min_row=2, max_row=2 + n)
    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + n)
    bar.add_data(dados, titles_from_data=True)
    bar.set_categories(cats)
    bar.height, bar.width = 10, 24

    linha = LineChart()
    linha.y_axis.axId = 200
    linha.y_axis.title = "% Acumulado"
    linha.y_axis.crosses = "max"
    dados_linha = Reference(ws, min_col=3, min_row=2, max_row=2 + n)
    linha.add_data(dados_linha, titles_from_data=True)
    bar += linha
    ws.add_chart(bar, "A35")

    bar2 = BarChart()
    bar2.title = "Pareto - Quantidade - Top 30 SKUs"
    bar2.y_axis.title = "Quantidade"
    bar2.x_axis.title = "Produto"
    dados2 = Reference(ws, min_col=7, min_row=2, max_row=2 + n)
    cats2 = Reference(ws, min_col=6, min_row=3, max_row=2 + n)
    bar2.add_data(dados2, titles_from_data=True)
    bar2.set_categories(cats2)
    bar2.height, bar2.width = 10, 24

    linha2 = LineChart()
    linha2.y_axis.axId = 200
    linha2.y_axis.title = "% Acumulado"
    linha2.y_axis.crosses = "max"
    dados_linha2 = Reference(ws, min_col=8, min_row=2, max_row=2 + n)
    linha2.add_data(dados_linha2, titles_from_data=True)
    bar2 += linha2
    ws.add_chart(bar2, "F35")

    autoajustar_colunas(ws)
    return ws


def gerar_planilha(resultado, resumo, resumo_cat):
    wb = Workbook()
    wb.remove(wb.active)
    aba_resumo(wb, resumo)
    aba_curva_abc(wb, resultado)
    aba_categoria(wb, resumo_cat)
    aba_graficos(wb, resultado)
    wb.save(ARQUIVO_SAIDA)


def main():
    print(f"Lendo: {ARQUIVO_ENTRADA.name}")
    df = carregar_dados()
    print(f"{len(df)} produtos validos carregados.")

    resultado = montar_tabela(df)
    resumo = montar_resumo(resultado)
    resumo_cat = montar_resumo_categoria(resultado)

    gerar_planilha(resultado, resumo, resumo_cat)
    print(f"Arquivo gerado: {ARQUIVO_SAIDA.name}")

    for _, r in resumo.iterrows():
        print(f"  [{r['Base']}] Classe {r['Classe']}: {int(r['Qtd_SKUs'])} SKUs "
              f"({r['Pct_SKUs']:.1%}) = {r['Pct_Metrica']:.1%} do total")


if __name__ == "__main__":
    main()
