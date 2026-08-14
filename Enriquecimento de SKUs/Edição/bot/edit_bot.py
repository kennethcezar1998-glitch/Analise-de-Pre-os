"""
Bot de automação de EDIÇÃO de Enriquecimento de SKU - Sportbay Marketplace.

Lê a planilha "Tags para enriquecimento - edição.xlsx" (aba "Tags"), loga no
backoffice, vai em OPERAÇÃO > Enriquecimento de SKU > Lista, busca cada SKU
(coluna A) na busca "Buscar por SKU", abre o registro já existente pelo botão
"Editar" e adiciona uma Tag Global para cada linha do grupo antes de salvar.

Diferente do bot de criação (pasta "bot" um nível acima, usado em
OPERAÇÃO > Enriquecimento de SKU > Criar), este bot NUNCA:
  - preenche "Código SKU da Sportbay" / "Código SKU do seller" (não editáveis
    na tela de edição);
  - altera "Curva ABC";
  - altera "Categorias predefinidas" (a coluna existe na planilha só como
    referência visual, não é usada aqui).

Ele só adiciona novas Tags Globais aos registros já cadastrados.

Uso:
    python edit_bot.py                 # roda tudo, navegador visível
    python edit_bot.py --dry-run       # preenche tudo mas NUNCA clica em Salvar
    python edit_bot.py --headless      # roda sem abrir janela do navegador
    python edit_bot.py --limit 1       # processa só o primeiro grupo (teste)
    python edit_bot.py --start-from 3  # pula os 2 primeiros grupos

Configuração (arquivo .env na mesma pasta):
    SPORTBAY_USER=usuario
    SPORTBAY_PASS=senha
    SPREADSHEET_PATH=caminho para o .xlsx (opcional, default abaixo)
"""

from __future__ import annotations

import argparse
import difflib
import logging
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urljoin

import openpyxl
from dotenv import load_dotenv
from playwright.sync_api import Locator, Page, TimeoutError as PWTimeoutError, sync_playwright

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
LOGS_DIR = BASE_DIR / "logs"
LOGS_DIR.mkdir(exist_ok=True)

load_dotenv(BASE_DIR / ".env")

SITE_URL = "https://www.marketplace.sportbay.com.br"
LOGIN_URL = f"{SITE_URL}/auth/jwt/login?returnTo=%2Fdashboard%2Fauction%2Flist"
LIST_URL = f"{SITE_URL}/dashboard/sku-enrichment"

USERNAME = os.getenv("SPORTBAY_USER")
PASSWORD = os.getenv("SPORTBAY_PASS")
SPREADSHEET_PATH = Path(
    os.getenv("SPREADSHEET_PATH", PROJECT_DIR / "Tags para enriquecimento - edição.xlsx")
)

# Confiança mínima (0-1) para aceitar uma opção de "Tag global" como correspondente
# às colunas E/F/G. Abaixo disso a tag é pulada e reportada para revisão manual.
TAG_MATCH_THRESHOLD = 0.55

DEFAULT_TIMEOUT = 15_000
SEARCH_DEBOUNCE_MS = 800

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOGS_DIR / "run.log", encoding="utf-8"),
    ],
)
log = logging.getLogger("edit_bot")


# --------------------------------------------------------------------------
# Planilha
# --------------------------------------------------------------------------

@dataclass
class TagRow:
    sku: str
    sku_seller: str
    grupo: str | None
    facet: str | None
    spec: str | None
    excel_row: int


@dataclass
class SkuGroup:
    sku: str
    sku_seller: str
    rows: list[TagRow]


def normalize(text) -> str:
    if text is None:
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return " ".join(text.split())


def load_tag_rows(path: Path) -> list[TagRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Tags"]
    rows: list[TagRow] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sku, sku_seller, grupo, facet, spec = row[0], row[1], row[4], row[5], row[6]
        if sku is None or sku_seller is None:
            continue
        rows.append(
            TagRow(
                sku=str(sku).strip(),
                sku_seller=str(sku_seller).strip(),
                grupo=str(grupo).strip() if grupo not in (None, "") else None,
                facet=str(facet).strip() if facet not in (None, "") else None,
                spec=str(spec).strip() if spec not in (None, "") else None,
                excel_row=i,
            )
        )
    return rows


def load_grupo_lookup(path: Path) -> dict[str, str]:
    """Mapa normalize(Rótulo) -> Código (data-value) a partir da aba Listas."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Listas"]
    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rotulo, codigo = row[2], row[3]
        if rotulo and codigo:
            lookup[normalize(rotulo)] = str(codigo).strip()
    return lookup


def group_consecutive(rows: list[TagRow]) -> list[SkuGroup]:
    """Agrupa linhas consecutivas que compartilham o mesmo par (SKU, SKU Seller).

    Cada grupo corresponde a UM registro já existente: o bot abre a tela de
    edição dele uma vez e adiciona uma Tag Global para cada linha do grupo,
    só clicando em Salvar depois da última linha.
    """
    groups: list[SkuGroup] = []
    for r in rows:
        if groups and groups[-1].sku == r.sku and groups[-1].sku_seller == r.sku_seller:
            groups[-1].rows.append(r)
        else:
            groups.append(SkuGroup(sku=r.sku, sku_seller=r.sku_seller, rows=[r]))
    return groups


# --------------------------------------------------------------------------
# Helpers de UI (MUI Select / MUI Autocomplete) - mesmos do bot de criação
# --------------------------------------------------------------------------

def best_fuzzy_match(query: str, candidates: list[str]) -> tuple[int, float]:
    """Retorna (índice, score 0-1) do candidato mais parecido com query."""
    q = normalize(query)
    best_idx, best_score = -1, 0.0
    for i, c in enumerate(candidates):
        score = difflib.SequenceMatcher(None, q, normalize(c)).ratio()
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx, best_score


def open_mui_select(page: Page, label_text: str) -> Locator:
    """Abre um MUI Select (native) pelo texto do label e retorna o listbox aberto."""
    trigger = page.get_by_role("button", name=label_text, exact=False)
    trigger.click()
    listbox = page.locator('ul[role="listbox"]').last
    listbox.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
    return listbox


def select_mui_option_by_data_value(page: Page, label_text: str, data_value: str) -> bool:
    listbox = open_mui_select(page, label_text)
    option = listbox.locator(f'li[role="option"][data-value="{data_value}"]')
    if option.count() == 0:
        page.keyboard.press("Escape")
        return False
    option.first.click()
    return True


def select_mui_option_by_text(page: Page, label_text: str, target_text: str, threshold: float) -> bool:
    listbox = open_mui_select(page, label_text)
    options = listbox.locator('li[role="option"]')
    texts = options.all_inner_texts()
    idx, score = best_fuzzy_match(target_text, texts)
    if idx == -1 or score < threshold:
        page.keyboard.press("Escape")
        return False
    options.nth(idx).click()
    return True


def autocomplete_select(
    page: Page,
    label_text: str,
    search_text: str,
    target_text: str,
    threshold: float,
) -> tuple[bool, float, str | None]:
    """Digita em um MUI Autocomplete e clica na opção mais parecida com target_text.

    Retorna (sucesso, score, texto_escolhido).
    """
    field = page.get_by_role("combobox", name=label_text, exact=False)
    field.click()
    field.fill("")
    field.type(search_text, delay=25)

    try:
        listbox = page.locator('ul[role="listbox"]').last
        listbox.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
    except PWTimeoutError:
        page.keyboard.press("Escape")
        return False, 0.0, None

    options = listbox.locator('li[role="option"]')
    count = options.count()
    if count == 0:
        page.keyboard.press("Escape")
        return False, 0.0, None

    texts = options.all_inner_texts()
    idx, score = best_fuzzy_match(target_text, texts)
    if idx == -1 or score < threshold:
        page.keyboard.press("Escape")
        return False, score, None

    chosen_text = texts[idx]
    options.nth(idx).click()
    return True, score, chosen_text


# --------------------------------------------------------------------------
# Fluxo principal
# --------------------------------------------------------------------------

def login(page: Page) -> None:
    log.info("Abrindo página de login...")
    page.goto(LOGIN_URL)
    page.fill('input[name="email"]', USERNAME)
    page.fill('input[name="password"]', PASSWORD)
    page.click('button[type="submit"]')
    page.wait_for_url("**/dashboard/**", timeout=30_000)
    log.info("Login OK, sessão iniciada.")


def find_edit_url(page: Page, sku: str, sku_seller: str) -> str | None:
    """Vai em OPERAÇÃO > Enriquecimento de SKU > Lista, busca pelo SKU
    (campo "Buscar por SKU") e retorna a URL de edição da linha cujo
    SKU Sportbay + SKU seller batem EXATAMENTE com os da planilha.

    Sem apertar Enter a busca não é disparada e a tabela fica com o
    resultado anterior (foi o que causou, num teste, a tabela mostrar
    3 SKUs diferentes em vez de só o buscado) - por isso, mesmo com o
    Enter, a função confere os valores da linha e nunca confia
    "no primeiro resultado" às cegas.
    """
    page.goto(LIST_URL)
    search = page.get_by_placeholder("Buscar por SKU")
    search.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
    search.click()
    search.fill("")
    search.type(sku, delay=30)
    search.press("Enter")
    page.wait_for_timeout(SEARCH_DEBOUNCE_MS)

    rows = page.locator("table tbody tr")
    try:
        rows.first.wait_for(state="visible", timeout=DEFAULT_TIMEOUT)
    except PWTimeoutError:
        return None

    count = rows.count()
    if count == 0:
        return None
    if count > 1:
        log.info(f'  {count} resultado(s) para SKU "{sku}" - procurando a linha com SKU/Seller exatos.')

    matches = []
    for i in range(count):
        row = rows.nth(i)
        textos = row.locator("a").all_inner_texts()
        if len(textos) < 2:
            continue
        row_sku, row_seller = textos[0].strip(), textos[1].strip()
        if row_sku == sku and row_seller == sku_seller:
            matches.append(row)

    if not matches:
        log.warning(f'  Nenhuma linha com SKU Sportbay="{sku}" e SKU Seller="{sku_seller}" exatos '
                    f'entre os {count} resultado(s) da busca.')
        return None
    if len(matches) > 1:
        log.warning(f'  {len(matches)} linhas batem exatamente com SKU/Seller "{sku}"/"{sku_seller}" '
                    f'- usando a primeira (verifique duplicidade).')

    editar = matches[0].get_by_role("link", name="Editar")
    href = editar.get_attribute("href")
    if not href:
        return None
    return urljoin(SITE_URL, href)


def add_tag_global(page: Page, row: TagRow, grupo_lookup: dict[str, str], pendencias: list[dict]) -> None:
    if not row.grupo:
        log.warning(f"  Linha Excel {row.excel_row}: sem Grupo (coluna E) - pulando tag")
        return

    codigo_grupo = grupo_lookup.get(normalize(row.grupo))
    opened = False
    if codigo_grupo:
        opened = select_mui_option_by_data_value(page, "Filtrar por grupo", codigo_grupo)
    if not opened:
        opened = select_mui_option_by_text(page, "Filtrar por grupo", row.grupo, threshold=0.6)
    if not opened:
        log.warning(f'  Linha Excel {row.excel_row}: Grupo "{row.grupo}" não encontrado no filtro - pulando tag')
        pendencias.append({"row": row, "motivo": "grupo não encontrado", "score": 0.0})
        return

    alvo = " ".join(p for p in (row.grupo, row.facet, row.spec) if p)
    busca = row.facet or row.grupo
    ok, score, chosen = autocomplete_select(
        page,
        label_text="Tag global",
        search_text=busca,
        target_text=alvo,
        threshold=TAG_MATCH_THRESHOLD,
    )
    if ok:
        log.info(f'  Tag "{row.grupo}" / "{row.facet}" -> "{chosen}" (score {score:.2f})')
    else:
        log.warning(f'  Linha Excel {row.excel_row}: nenhuma Tag global bateu com confiança suficiente '
                     f'(melhor score {score:.2f}) para "{alvo}" - pulei, revisar manualmente')
        pendencias.append({"row": row, "motivo": "tag sem match confiável", "score": score})


def confirm_edit_page_sku(page: Page, sku: str, sku_seller: str) -> bool:
    """Confere, na própria tela de edição já aberta, se os campos (bloqueados)
    "Código SKU da SportBay" e "Código SKU do seller" batem com a planilha.

    Segunda camada de segurança além do match feito na tela de Lista - só
    prossegue com as Tags Globais se os dois valores exibidos forem iguais.
    """
    try:
        campo_sku = page.get_by_label("Código SKU da SportBay", exact=False)
        campo_seller = page.get_by_label("Código SKU do seller", exact=False)
        valor_sku = campo_sku.input_value(timeout=DEFAULT_TIMEOUT).strip()
        valor_seller = campo_seller.input_value(timeout=DEFAULT_TIMEOUT).strip()
    except PWTimeoutError:
        log.error("  Não consegui ler os campos de SKU da tela de edição para conferir.")
        return False
    return valor_sku == sku and valor_seller == sku_seller


def wait_salvar_habilitado(page: Page, timeout_ms: int = DEFAULT_TIMEOUT) -> Locator:
    botao = page.get_by_role("button", name="Atualizar Enriquecimento")
    botao.wait_for(state="visible", timeout=timeout_ms)
    page.wait_for_function(
        """(btn) => btn && !btn.disabled && !btn.classList.contains('Mui-disabled')""",
        arg=botao.element_handle(),
        timeout=timeout_ms,
    )
    return botao


def process_group(page: Page, group: SkuGroup, grupo_lookup: dict[str, str],
                   idx: int, total: int, dry_run: bool, pendencias: list[dict],
                   nao_encontrados: list[SkuGroup]) -> bool:
    log.info(f"[{idx}/{total}] SKU Sportbay={group.sku} | SKU Seller={group.sku_seller} "
             f"({len(group.rows)} tag(s) novas)")

    edit_url = find_edit_url(page, group.sku, group.sku_seller)
    if not edit_url:
        log.warning(f'  SKU "{group.sku}"/"{group.sku_seller}" não encontrado na Lista - '
                    f'pulando (verifique se já está cadastrado)')
        nao_encontrados.append(group)
        return False

    page.goto(edit_url)
    # Não mexe em "Código SKU da Sportbay", "Código SKU do seller" (não editáveis)
    # nem em "Curva ABC" - só adiciona Tags Globais.
    page.get_by_role("combobox", name="Tag global").wait_for(state="visible", timeout=DEFAULT_TIMEOUT)

    if not confirm_edit_page_sku(page, group.sku, group.sku_seller):
        log.error(f'  Tela de edição aberta ({edit_url}) não confere com SKU "{group.sku}"/"{group.sku_seller}" '
                  f'- abortando este grupo para não editar o registro errado.')
        page.screenshot(path=str(LOGS_DIR / f"erro_grupo_{idx}_sku_nao_confere.png"))
        nao_encontrados.append(group)
        return False

    for row in group.rows:
        add_tag_global(page, row, grupo_lookup, pendencias)

    if dry_run:
        screenshot_path = LOGS_DIR / f"dry_run_grupo_{idx}.png"
        page.screenshot(path=str(screenshot_path), full_page=True)
        log.info(f"  [DRY-RUN] não vou clicar em Salvar. Screenshot: {screenshot_path}")
        return True

    try:
        botao = wait_salvar_habilitado(page)
    except PWTimeoutError:
        log.error("  Botão 'Atualizar Enriquecimento' não habilitou - screenshot salvo, pulando grupo")
        page.screenshot(path=str(LOGS_DIR / f"erro_grupo_{idx}_botao_desabilitado.png"))
        return False

    botao.click()
    try:
        page.wait_for_url("**/dashboard/sku-enrichment**", timeout=DEFAULT_TIMEOUT)
        log.info("  Salvo com sucesso.")
    except PWTimeoutError:
        log.warning("  Não confirmei o salvamento automaticamente - verifique manualmente.")
        page.screenshot(path=str(LOGS_DIR / f"aviso_grupo_{idx}_confirmar_salvamento.png"))
    return True


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="preenche tudo mas não clica em Salvar")
    parser.add_argument("--headless", action="store_true", help="roda sem interface gráfica")
    parser.add_argument("--limit", type=int, default=None, help="processa só os N primeiros grupos")
    parser.add_argument("--start-from", type=int, default=1, help="índice (1-based) do grupo inicial")
    args = parser.parse_args()

    if not USERNAME or not PASSWORD:
        log.error('Defina SPORTBAY_USER e SPORTBAY_PASS no arquivo "bot/.env" antes de rodar.')
        sys.exit(1)
    if not SPREADSHEET_PATH.exists():
        log.error(f"Planilha não encontrada: {SPREADSHEET_PATH}")
        sys.exit(1)

    rows = load_tag_rows(SPREADSHEET_PATH)
    grupo_lookup = load_grupo_lookup(SPREADSHEET_PATH)
    groups = group_consecutive(rows)
    log.info(f"{len(rows)} linha(s) de tag lidas -> {len(groups)} registro(s) a editar.")

    start = max(args.start_from - 1, 0)
    end = start + args.limit if args.limit else len(groups)
    groups = groups[start:end]

    pendencias: list[dict] = []
    nao_encontrados: list[SkuGroup] = []
    sucesso, falha = 0, 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless, slow_mo=0 if args.headless else 150)
        context = browser.new_context()
        page = context.new_page()
        page.set_default_timeout(DEFAULT_TIMEOUT)

        login(page)

        for i, group in enumerate(groups, start=start + 1):
            try:
                ok = process_group(page, group, grupo_lookup, i, start + len(groups), args.dry_run,
                                    pendencias, nao_encontrados)
                sucesso += 1 if ok else 0
                falha += 0 if ok else 1
            except Exception:
                log.exception(f"Erro inesperado no grupo {i} (SKU {group.sku})")
                page.screenshot(path=str(LOGS_DIR / f"erro_grupo_{i}_excecao.png"))
                falha += 1

        browser.close()

    log.info(f"Concluído: {sucesso} registro(s) ok, {falha} registro(s) com falha.")
    if nao_encontrados:
        log.warning(f"{len(nao_encontrados)} SKU(s) não encontrados na Lista:")
        for g in nao_encontrados:
            log.warning(f"  SKU {g.sku}/{g.sku_seller}")
    if pendencias:
        log.warning(f"{len(pendencias)} tag(s) pulada(s) por falta de correspondência confiável:")
        for p in pendencias:
            r = p["row"]
            log.warning(f'  Excel linha {r.excel_row}: SKU {r.sku}/{r.sku_seller} '
                        f'Grupo="{r.grupo}" Facet="{r.facet}" - motivo: {p["motivo"]} (score {p["score"]:.2f})')


if __name__ == "__main__":
    main()
